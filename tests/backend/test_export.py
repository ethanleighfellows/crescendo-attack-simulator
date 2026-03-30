"""Unit tests for XLSX export generation.

Verifies sheet structure, column headers, data population,
and redaction behavior.
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook

from src.backend.engine.types import (
    JudgeDecision,
    Outcome,
    RunConfig,
    RunResult,
    RunStatus,
    TurnRecord,
)
from src.backend.export.schemas import (
    CONFIG_COLUMNS,
    OUTCOMES_COLUMNS,
    SUMMARY_COLUMNS,
    TRANSCRIPT_COLUMNS,
    ExportConfig,
)
from src.backend.export.xlsx import export_to_xlsx


def _make_result() -> RunResult:
    now = datetime.utcnow()
    return RunResult(
        run_id="test-run-123",
        config=RunConfig(
            objective="Test objective",
            vulnerability="Test vuln",
            vulnerability_type="Test type",
            max_rounds=5,
            max_backtracks=3,
            simulator_api_key="sk-test12345678901234567890",
        ),
        status=RunStatus.COMPLETED,
        outcome=Outcome.SUCCESS,
        turns=[
            TurnRecord(
                round_number=1,
                backtrack_count=0,
                user_prompt="Tell me about history",
                assistant_response="Here is some history...",
                turn_level_attack=None,
                enhancement_applied=False,
                refusal_decision=JudgeDecision(
                    value=False, rationale="Not a refusal", metadata=0
                ),
                eval_decision=JudgeDecision(
                    value=False, rationale="Partial match", metadata=40
                ),
                is_backtrack=False,
                timestamp=now,
            ),
            TurnRecord(
                round_number=2,
                backtrack_count=0,
                user_prompt="Now tell me the details",
                assistant_response="Here are the specific details...",
                turn_level_attack="base64",
                enhancement_applied=True,
                refusal_decision=JudgeDecision(
                    value=False, rationale="Not a refusal", metadata=0
                ),
                eval_decision=JudgeDecision(
                    value=True, rationale="Full match", metadata=100
                ),
                is_backtrack=False,
                timestamp=now + timedelta(seconds=30),
            ),
        ],
        total_rounds=2,
        total_backtracks=0,
        final_eval_score=100,
        started_at=now,
        completed_at=now + timedelta(seconds=60),
    )


class TestXlsxExport:
    def test_produces_valid_xlsx_bytes(self) -> None:
        result = _make_result()
        config = ExportConfig()
        xlsx_bytes = export_to_xlsx(result, config)
        assert len(xlsx_bytes) > 0
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert wb is not None

    def test_has_four_sheets(self) -> None:
        result = _make_result()
        config = ExportConfig()
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, config)))
        sheet_names = wb.sheetnames
        assert "Run Summary" in sheet_names
        assert "Configuration" in sheet_names
        assert "Full Transcript" in sheet_names
        assert "Per-Turn Outcomes" in sheet_names

    def test_summary_sheet_headers(self) -> None:
        result = _make_result()
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, ExportConfig())))
        ws = wb["Run Summary"]
        headers = [cell.value for cell in ws[1]]
        assert headers == SUMMARY_COLUMNS

    def test_config_sheet_headers(self) -> None:
        result = _make_result()
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, ExportConfig())))
        ws = wb["Configuration"]
        headers = [cell.value for cell in ws[1]]
        assert headers == CONFIG_COLUMNS

    def test_transcript_sheet_has_correct_row_count(self) -> None:
        result = _make_result()
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, ExportConfig())))
        ws = wb["Full Transcript"]
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        assert len(data_rows) == 2

    def test_outcomes_sheet_headers(self) -> None:
        result = _make_result()
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, ExportConfig())))
        ws = wb["Per-Turn Outcomes"]
        headers = [cell.value for cell in ws[1]]
        assert headers == OUTCOMES_COLUMNS

    def test_api_keys_excluded_by_default(self) -> None:
        result = _make_result()
        config = ExportConfig(include_api_keys=False)
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, config)))
        ws = wb["Configuration"]
        values = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert "Simulator API Key" not in values

    def test_api_keys_included_when_requested(self) -> None:
        result = _make_result()
        config = ExportConfig(include_api_keys=True)
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, config)))
        ws = wb["Configuration"]
        values = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert "Simulator API Key" in values

    def test_redaction_masks_api_keys(self) -> None:
        result = _make_result()
        result.turns[0].user_prompt = "Use key sk-abcdefghijklmnopqrstuvwxyz123456"
        config = ExportConfig(redact_sensitive_content=True)
        wb = load_workbook(io.BytesIO(export_to_xlsx(result, config)))
        ws = wb["Full Transcript"]
        user_prompt_cell = ws.cell(row=2, column=3).value
        assert "sk-" not in str(user_prompt_cell)
        assert "[REDACTED]" in str(user_prompt_cell)
