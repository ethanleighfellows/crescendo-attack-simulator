from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.backend.engine.types import RunResult
from src.backend.export.schemas import (
    CONFIG_COLUMNS,
    OUTCOMES_COLUMNS,
    SUMMARY_COLUMNS,
    TRANSCRIPT_COLUMNS,
    ExportConfig,
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

REDACTION_PATTERN = re.compile(
    r"(sk-[a-zA-Z0-9]{20,}|[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})",
    re.IGNORECASE,
)


def _redact(text: str) -> str:
    return REDACTION_PATTERN.sub("[REDACTED]", text)


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _auto_width(ws: Any, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_len = max(max_len, min(cell_len, max_width))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4


def export_to_xlsx(result: RunResult, config: ExportConfig) -> bytes:
    """Generate a multi-sheet XLSX workbook from a RunResult.

    Sheets:
      1. Run Summary — high-level outcome and timing.
      2. Configuration — all run parameters.
      3. Full Transcript — one row per turn with prompts and responses.
      4. Per-Turn Outcomes — judge decisions and scores per round.

    Returns:
        The XLSX file as bytes.
    """
    wb = Workbook()
    redact = _redact if config.redact_sensitive_content else lambda t: t

    _write_summary_sheet(wb, result, redact)
    _write_config_sheet(wb, result, config)
    _write_transcript_sheet(wb, result, redact)
    _write_outcomes_sheet(wb, result, config, redact)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_summary_sheet(
    wb: Workbook, result: RunResult, redact: Any
) -> None:
    ws = wb.active
    ws.title = "Run Summary"
    ws.append(SUMMARY_COLUMNS)
    _style_header(ws)

    duration = None
    if result.started_at and result.completed_at:
        duration = (result.completed_at - result.started_at).total_seconds()

    ws.append([
        result.run_id,
        redact(result.config.objective),
        result.config.vulnerability,
        result.config.vulnerability_type,
        result.outcome.value if result.outcome else "",
        result.total_rounds,
        result.total_backtracks,
        result.final_eval_score,
        str(result.started_at) if result.started_at else "",
        str(result.completed_at) if result.completed_at else "",
        duration,
    ])
    _auto_width(ws)


def _write_config_sheet(
    wb: Workbook, result: RunResult, export_config: ExportConfig
) -> None:
    ws = wb.create_sheet("Configuration")
    ws.append(CONFIG_COLUMNS)
    _style_header(ws)

    cfg = result.config
    rows = [
        ("Objective", cfg.objective),
        ("Vulnerability", cfg.vulnerability),
        ("Vulnerability Type", cfg.vulnerability_type),
        ("Max Rounds", cfg.max_rounds),
        ("Max Backtracks", cfg.max_backtracks),
        ("Enhancement Probability", cfg.enhancement_probability),
        ("Success Threshold", cfg.success_threshold),
        ("Stop on First Success", cfg.stop_on_first_success),
        ("Turn-Level Attacks", ", ".join(cfg.turn_level_attacks) if cfg.turn_level_attacks else "None"),
        ("Simulator Provider", cfg.simulator_provider),
        ("Simulator Model", cfg.simulator_model),
        ("Target Provider", cfg.target_provider),
        ("Target Model", cfg.target_model),
        ("Temperature", cfg.temperature),
        ("Max Retries", cfg.max_retries),
    ]

    if export_config.include_api_keys:
        rows.append(("Simulator API Key", cfg.simulator_api_key or ""))
        rows.append(("Target API Key", cfg.target_api_key or ""))

    for param, value in rows:
        ws.append([param, str(value)])
    _auto_width(ws)


def _write_transcript_sheet(
    wb: Workbook, result: RunResult, redact: Any
) -> None:
    ws = wb.create_sheet("Full Transcript")
    ws.append(TRANSCRIPT_COLUMNS)
    _style_header(ws)

    for turn in result.turns:
        ws.append([
            turn.round_number,
            turn.backtrack_count,
            redact(turn.user_prompt),
            redact(turn.assistant_response),
            turn.turn_level_attack or "",
            "Yes" if turn.enhancement_applied else "No",
            "Yes" if turn.is_backtrack else "No",
            str(turn.timestamp),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    _auto_width(ws)


def _write_outcomes_sheet(
    wb: Workbook,
    result: RunResult,
    export_config: ExportConfig,
    redact: Any,
) -> None:
    ws = wb.create_sheet("Per-Turn Outcomes")
    ws.append(OUTCOMES_COLUMNS)
    _style_header(ws)

    for turn in result.turns:
        refusal_yn = ""
        refusal_rationale = ""
        refusal_score = ""
        if turn.refusal_decision:
            refusal_yn = "Yes" if turn.refusal_decision.value else "No"
            refusal_score = turn.refusal_decision.metadata
            if export_config.include_judge_rationale:
                refusal_rationale = redact(turn.refusal_decision.rationale)

        eval_yn = ""
        eval_score = ""
        eval_rationale = ""
        if turn.eval_decision:
            eval_yn = "Yes" if turn.eval_decision.value else "No"
            eval_score = turn.eval_decision.metadata
            if export_config.include_judge_rationale:
                eval_rationale = redact(turn.eval_decision.rationale)

        ws.append([
            turn.round_number,
            turn.backtrack_count,
            refusal_yn,
            refusal_rationale,
            refusal_score,
            eval_yn,
            eval_score,
            eval_rationale,
            turn.turn_level_attack or "",
        ])

    _auto_width(ws)
